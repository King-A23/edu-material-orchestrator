from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ...models.source import SourceUnit
from ...utils.hashing import make_manifest_id
from ...utils.provenance import make_source_ref
from ..common.base import DetectedInput, ReadError, ReaderOutput, SourceReader, build_source_document


def _sheet_preview(sheet, max_rows: int = 20, max_cols: int = 10) -> str:
    lines = [f"Worksheet: {sheet.title}", f"Dimensions: {sheet.max_row} x {sheet.max_column}"]
    for row in sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if values:
            lines.append("\t".join(values))
    return "\n".join(lines)


class XlsxReader(SourceReader):
    supported_type = "xlsx"

    def read(self, detected: DetectedInput) -> ReaderOutput:
        workbook_path = Path(detected.resolved_path)
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as error:
            raise ReadError(f"Failed to open XLSX file '{workbook_path}': {error}") from error

        document = build_source_document(detected)
        units: list[SourceUnit] = []

        try:
            for index, sheet_name in enumerate(workbook.sheetnames, start=1):
                sheet = workbook[sheet_name]
                source_ref = make_source_ref("xlsx", index)
                units.append(
                    SourceUnit(
                        unit_id=make_manifest_id(document.id, source_ref),
                        source_id=document.id,
                        index=index,
                        kind="sheet",
                        raw_text=_sheet_preview(sheet),
                        confidence=1.0,
                        source_ref=source_ref,
                    )
                )
        finally:
            workbook.close()

        return ReaderOutput(
            document=document,
            units=units,
            metadata={
                "reader": "open.xlsx_reader",
                "sheet_count": len(units),
            },
        )
