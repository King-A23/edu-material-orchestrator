from __future__ import annotations

import mimetypes
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pptx import Presentation
from pypdf import PdfReader

from ..backends.common.base import DetectedInput, ReadError
from ..utils.files import guess_file_type, normalize_input_path
from ..utils.hashing import hash_file


MIME_OVERRIDES = {
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "pdf": "application/pdf",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _count_pptx_slides(path: Path) -> int:
    presentation = Presentation(path)
    return len(presentation.slides)


def _count_pdf_pages(path: Path) -> int:
    reader = PdfReader(str(path))
    return len(reader.pages)


def _count_docx_paragraphs(path: Path) -> int:
    document = Document(path)
    return len(document.paragraphs)


def _count_xlsx_sheets(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return len(workbook.sheetnames)
    finally:
        workbook.close()


def count_units(path: Path, file_type: str) -> int | None:
    try:
        if file_type == "pptx":
            return _count_pptx_slides(path)
        if file_type == "pdf":
            return _count_pdf_pages(path)
        if file_type == "docx":
            return _count_docx_paragraphs(path)
        if file_type == "xlsx":
            return _count_xlsx_sheets(path)
    except Exception as error:
        raise ReadError(f"Failed to inspect '{path}': {error}") from error
    return None


def detect_input(path: str | Path) -> DetectedInput:
    resolved = normalize_input_path(path)
    file_type = guess_file_type(resolved)
    mime_type = MIME_OVERRIDES.get(file_type) or mimetypes.guess_type(str(resolved))[0]
    return DetectedInput(
        path=str(path),
        resolved_path=str(resolved),
        type=file_type,
        title=resolved.stem,
        mime_type=mime_type,
        page_or_slide_count=count_units(resolved, file_type),
        source_hash=hash_file(resolved),
        size_bytes=resolved.stat().st_size,
    )
